from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Count
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from datetime import date, timedelta
import json

from .models import Employe, DemandeConge, TypeConge, SoldeConge, Notification, Departement


def connexion(request):
    if request.user.is_authenticated:
        return redirect('tableau_de_bord')
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        try:
            user_obj = Employe.objects.get(email=email)
            user = authenticate(request, username=user_obj.username, password=password)
            if user and user.actif:
                login(request, user)
                next_url = request.GET.get('next', 'tableau_de_bord')
                return redirect(next_url)
        except Employe.DoesNotExist:
            pass
        messages.error(request, 'Email ou mot de passe incorrect.')
    return render(request, 'conges/connexion.html')


def deconnexion(request):
    logout(request)
    return redirect('connexion')


@login_required
def tableau_de_bord(request):
    user = request.user
    annee = date.today().year

    mes_demandes_recentes = DemandeConge.objects.filter(
        employe=user
    ).select_related('type_conge')[:5]

    mes_demandes_en_attente = DemandeConge.objects.filter(
        employe=user, statut='en_attente'
    ).count()

    mes_demandes_approuvees = DemandeConge.objects.filter(
        employe=user, statut='approuve',
        date_soumission__year=annee
    ).count()

    mes_soldes = SoldeConge.objects.filter(
        employe=user, annee=annee
    ).select_related('type_conge')

    a_approuver = 0
    demandes_equipe = []
    if user.is_manager_or_above:
        if user.role == 'manager':
            equipe_ids = user.subordonnes.values_list('id', flat=True)
            a_approuver = DemandeConge.objects.filter(
                employe__in=equipe_ids, statut='en_attente'
            ).count()
        elif user.role in ('rh', 'admin'):
            a_approuver = DemandeConge.objects.filter(statut='en_attente').count()

        demandes_equipe = DemandeConge.objects.filter(
            statut='approuve',
            date_fin__gte=date.today()
        ).select_related('employe', 'type_conge')[:10]

    notifications_non_lues = user.notifications.filter(lue=False).count()

    context = {
        'mes_demandes_recentes': mes_demandes_recentes,
        'mes_demandes_en_attente': mes_demandes_en_attente,
        'mes_demandes_approuvees': mes_demandes_approuvees,
        'mes_soldes': mes_soldes,
        'a_approuver': a_approuver,
        'demandes_equipe': demandes_equipe,
        'notifications_non_lues': notifications_non_lues,
        'today': date.today(),
    }
    return render(request, 'conges/tableau_de_bord.html', context)


@login_required
def nouvelle_demande(request):
    user = request.user
    types_conge = TypeConge.objects.filter(actif=True)
    annee = date.today().year
    soldes_qs = SoldeConge.objects.filter(employe=user, annee=annee)
    soldes = {s.type_conge_id: s for s in soldes_qs}
    soldes_json = {str(s.type_conge_id): float(s.jours_restants) for s in soldes_qs}
    employes_liste = Employe.objects.filter(actif=True).exclude(id=user.id).order_by('last_name', 'first_name')

    if request.method == 'POST':
        type_id = request.POST.get('type_conge')
        date_debut_str = request.POST.get('date_debut')
        date_fin_str = request.POST.get('date_fin')
        demi_journee = request.POST.get('demi_journee') == 'on'
        periode_demi_journee = request.POST.get('periode_demi_journee', '')
        motif = request.POST.get('motif', '').strip()
        justificatif = request.FILES.get('justificatif')
        interimaire_id = request.POST.get('interimaire') or None

        errors = []
        if not type_id:
            errors.append("Veuillez sélectionner un type de congé.")
        if not date_debut_str:
            errors.append("La date de début est requise.")
        if not date_fin_str:
            errors.append("La date de fin est requise.")

        type_conge_obj = None
        if type_id:
            try:
                type_conge_obj = TypeConge.objects.get(id=type_id)
                if type_conge_obj.categorie == 'absence' and not motif:
                    errors.append("Le motif est requis pour une demande d'absence.")
            except TypeConge.DoesNotExist:
                errors.append("Type invalide.")

        if not errors:
            try:
                type_conge = type_conge_obj or TypeConge.objects.get(id=type_id)
                date_debut = date.fromisoformat(date_debut_str)
                date_fin = date.fromisoformat(date_fin_str)

                if date_debut > date_fin:
                    errors.append("La date de début doit être avant la date de fin.")
                elif date_debut < date.today():
                    errors.append("La date de début ne peut pas être dans le passé.")
                else:
                    if type_conge.necessite_justificatif and not justificatif:
                        errors.append(f"Un justificatif est requis pour '{type_conge.libelle}'.")
                    else:
                        demande = DemandeConge(
                            employe=user,
                            type_conge=type_conge,
                            date_debut=date_debut,
                            date_fin=date_fin,
                            demi_journee=demi_journee,
                            periode_demi_journee=periode_demi_journee if demi_journee else '',
                            motif=motif,
                            interimaire_id=interimaire_id,
                        )
                        if justificatif:
                            demande.justificatif = justificatif
                        demande.save()

                        _notifier_manager(demande)

                        messages.success(request, f"Votre demande {demande.reference} a été soumise avec succès.")
                        return redirect('mes_demandes')
            except TypeConge.DoesNotExist:
                errors.append("Type de congé invalide.")
            except ValueError:
                errors.append("Format de date invalide.")

        for err in errors:
            messages.error(request, err)

    context = {
        'types_conge': types_conge,
        'soldes': soldes,
        'soldes_json': json.dumps(soldes_json),
        'employes_liste': employes_liste,
        'today': date.today().isoformat(),
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/nouvelle_demande.html', context)


def _notifier_manager(demande):
    employe = demande.employe
    if employe.manager:
        Notification.objects.create(
            destinataire=employe.manager,
            titre=f"Nouvelle demande de congé",
            message=f"{employe.get_full_name()} a soumis une demande de {demande.type_conge.libelle} "
                    f"du {demande.date_debut.strftime('%d/%m/%Y')} au {demande.date_fin.strftime('%d/%m/%Y')}.",
            lien=f"/approbations/{demande.id}/",
        )
    for rh in Employe.objects.filter(role__in=['rh', 'admin'], actif=True):
        Notification.objects.create(
            destinataire=rh,
            titre=f"Nouvelle demande - {employe.get_full_name()}",
            message=f"Demande {demande.reference} : {demande.type_conge.libelle} "
                    f"({demande.nombre_jours} jour(s)).",
            lien=f"/approbations/{demande.id}/",
        )


@login_required
def mes_demandes(request):
    user = request.user
    statut_filtre = request.GET.get('statut', '')
    type_filtre = request.GET.get('type', '')
    annee_filtre = request.GET.get('annee', str(date.today().year))

    demandes = DemandeConge.objects.filter(employe=user).select_related('type_conge', 'traite_par')

    if statut_filtre:
        demandes = demandes.filter(statut=statut_filtre)
    if type_filtre:
        demandes = demandes.filter(type_conge__code=type_filtre)
    if annee_filtre:
        demandes = demandes.filter(date_soumission__year=annee_filtre)

    types_conge = TypeConge.objects.filter(actif=True)
    annees = range(date.today().year, date.today().year - 4, -1)

    context = {
        'demandes': demandes,
        'types_conge': types_conge,
        'statut_filtre': statut_filtre,
        'type_filtre': type_filtre,
        'annee_filtre': annee_filtre,
        'annees': annees,
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/mes_demandes.html', context)


@login_required
def annuler_demande(request, demande_id):
    demande = get_object_or_404(DemandeConge, id=demande_id, employe=request.user)
    if demande.statut == 'en_attente':
        demande.statut = 'annule'
        demande.save()
        messages.success(request, f"La demande {demande.reference} a été annulée.")
    else:
        messages.error(request, "Cette demande ne peut plus être annulée.")
    return redirect('mes_demandes')


@login_required
def mon_solde(request):
    user = request.user
    annee = int(request.GET.get('annee', date.today().year))
    soldes = SoldeConge.objects.filter(
        employe=user, annee=annee
    ).select_related('type_conge').order_by('type_conge__libelle')

    demandes_approuvees = DemandeConge.objects.filter(
        employe=user,
        statut='approuve',
        date_debut__year=annee
    ).select_related('type_conge').order_by('-date_debut')

    annees = range(date.today().year, date.today().year - 3, -1)

    context = {
        'soldes': soldes,
        'demandes_approuvees': demandes_approuvees,
        'annee': annee,
        'annees': annees,
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/mon_solde.html', context)


@login_required
def calendrier(request):
    user = request.user
    today = date.today()
    mois = int(request.GET.get('mois', today.month))
    annee_cal = int(request.GET.get('annee', today.year))

    if mois < 1:
        mois = 12
        annee_cal -= 1
    elif mois > 12:
        mois = 1
        annee_cal += 1

    premier_jour = date(annee_cal, mois, 1)
    if mois == 12:
        dernier_jour = date(annee_cal + 1, 1, 1) - timedelta(days=1)
    else:
        dernier_jour = date(annee_cal, mois + 1, 1) - timedelta(days=1)

    if user.is_manager_or_above:
        demandes = DemandeConge.objects.filter(
            statut='approuve',
            date_debut__lte=dernier_jour,
            date_fin__gte=premier_jour
        ).select_related('employe', 'type_conge')
    else:
        equipe_ids = [user.id]
        if user.manager:
            equipe_ids += list(Employe.objects.filter(manager=user.manager).values_list('id', flat=True))
        demandes = DemandeConge.objects.filter(
            employe__in=equipe_ids,
            statut='approuve',
            date_debut__lte=dernier_jour,
            date_fin__gte=premier_jour
        ).select_related('employe', 'type_conge')

    evenements = []
    for d in demandes:
        evenements.append({
            'id': d.id,
            'titre': d.employe.get_full_name(),
            'type': d.type_conge.libelle,
            'couleur': d.type_conge.couleur,
            'debut': d.date_debut.isoformat(),
            'fin': d.date_fin.isoformat(),
            'jours': float(d.nombre_jours),
            'est_moi': d.employe_id == user.id,
        })

    mois_nom = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre'][mois]

    context = {
        'evenements_json': json.dumps(evenements),
        'mois': mois,
        'annee': annee_cal,
        'mois_nom': mois_nom,
        'premier_jour': premier_jour,
        'dernier_jour': dernier_jour,
        'today': today,
        'mois_precedent': (mois - 1) or 12,
        'annee_precedent': annee_cal if mois > 1 else annee_cal - 1,
        'mois_suivant': (mois % 12) + 1,
        'annee_suivant': annee_cal if mois < 12 else annee_cal + 1,
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/calendrier.html', context)


@login_required
def approbations(request):
    user = request.user
    if not user.is_manager_or_above:
        messages.error(request, "Accès non autorisé.")
        return redirect('tableau_de_bord')

    statut_filtre = request.GET.get('statut', 'en_attente')

    if user.role == 'manager':
        equipe_ids = user.subordonnes.values_list('id', flat=True)
        demandes = DemandeConge.objects.filter(
            employe__in=equipe_ids
        ).select_related('employe', 'type_conge', 'traite_par')
    else:
        demandes = DemandeConge.objects.all().select_related('employe', 'type_conge', 'traite_par')

    if statut_filtre:
        demandes = demandes.filter(statut=statut_filtre)

    stats = {
        'en_attente': demandes.filter(statut='en_attente').count() if not statut_filtre else
                      DemandeConge.objects.filter(statut='en_attente').count()
                      if user.role in ('rh', 'admin') else
                      DemandeConge.objects.filter(statut='en_attente', employe__in=user.subordonnes.all()).count(),
        'approuve': 0,
        'rejete': 0,
    }

    context = {
        'demandes': demandes,
        'statut_filtre': statut_filtre,
        'stats': stats,
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/approbations.html', context)


@login_required
def detail_demande(request, demande_id):
    user = request.user
    if user.is_manager_or_above:
        demande = get_object_or_404(DemandeConge, id=demande_id)
    else:
        demande = get_object_or_404(DemandeConge, id=demande_id, employe=user)

    context = {
        'demande': demande,
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/detail_demande.html', context)


@login_required
@require_POST
def traiter_demande(request, demande_id):
    user = request.user
    if not user.is_manager_or_above:
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    demande = get_object_or_404(DemandeConge, id=demande_id)

    if demande.statut != 'en_attente':
        messages.error(request, "Cette demande a déjà été traitée.")
        return redirect('approbations')

    action = request.POST.get('action')
    commentaire = request.POST.get('commentaire', '').strip()

    if action not in ('approuver', 'rejeter'):
        messages.error(request, "Action invalide.")
        return redirect('approbations')

    demande.statut = 'approuve' if action == 'approuver' else 'rejete'
    demande.date_traitement = timezone.now()
    demande.traite_par = user
    demande.commentaire_traitement = commentaire
    demande.save()

    if action == 'approuver':
        _mettre_a_jour_solde(demande)

    _notifier_employe(demande, action, commentaire)

    msg = "approuvée" if action == 'approuver' else "rejetée"
    messages.success(request, f"La demande {demande.reference} a été {msg}.")
    return redirect('approbations')


def _mettre_a_jour_solde(demande):
    annee = demande.date_debut.year
    solde, _ = SoldeConge.objects.get_or_create(
        employe=demande.employe,
        type_conge=demande.type_conge,
        annee=annee,
        defaults={'jours_acquis': demande.type_conge.jours_max}
    )
    solde.jours_pris += demande.nombre_jours
    solde.save()


def _notifier_employe(demande, action, commentaire):
    statut_label = "approuvée" if action == 'approuver' else "rejetée"
    msg = f"Votre demande {demande.reference} ({demande.type_conge.libelle}) a été {statut_label}."
    if commentaire:
        msg += f" Commentaire : {commentaire}"
    Notification.objects.create(
        destinataire=demande.employe,
        titre=f"Demande {statut_label}",
        message=msg,
        lien=f"/mes-demandes/",
    )


@login_required
def equipe(request):
    user = request.user
    if not user.is_manager_or_above:
        messages.error(request, "Accès non autorisé.")
        return redirect('tableau_de_bord')

    if user.role == 'manager':
        employes = Employe.objects.filter(manager=user, actif=True).select_related('departement')
    else:
        employes = Employe.objects.filter(actif=True).select_related('departement', 'manager')

    annee = date.today().year
    aujourd_hui = date.today()
    absents_aujourd_hui = DemandeConge.objects.filter(
        statut='approuve',
        date_debut__lte=aujourd_hui,
        date_fin__gte=aujourd_hui,
    ).select_related('employe', 'type_conge')

    if user.role == 'manager':
        equipe_ids = user.subordonnes.values_list('id', flat=True)
        absents_aujourd_hui = absents_aujourd_hui.filter(employe__in=equipe_ids)

    context = {
        'employes': employes,
        'absents_aujourd_hui': absents_aujourd_hui,
        'aujourd_hui': aujourd_hui,
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/equipe.html', context)


@login_required
def administration(request):
    user = request.user
    if user.role not in ('rh', 'admin'):
        messages.error(request, "Accès non autorisé.")
        return redirect('tableau_de_bord')

    onglet = request.GET.get('onglet', 'employes')
    employes = Employe.objects.filter(actif=True).select_related('departement').order_by('last_name')
    departements = Departement.objects.all()
    types_conge = TypeConge.objects.all()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add_employe':
            return _ajouter_employe(request)
        elif action == 'add_type_conge':
            return _ajouter_type_conge(request)
        elif action == 'init_soldes':
            return _initialiser_soldes(request)

    context = {
        'onglet': onglet,
        'employes': employes,
        'departements': departements,
        'types_conge': types_conge,
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/administration.html', context)


def _ajouter_employe(request):
    try:
        dept_id = request.POST.get('departement')
        manager_id = request.POST.get('manager')
        username = request.POST.get('email', '').split('@')[0]
        emp = Employe.objects.create_user(
            username=username,
            email=request.POST.get('email'),
            password=request.POST.get('password', 'Petrosen2025!'),
            first_name=request.POST.get('prenom', ''),
            last_name=request.POST.get('nom', ''),
        )
        emp.poste = request.POST.get('poste', '')
        emp.role = request.POST.get('role', 'employe')
        emp.matricule = request.POST.get('matricule', '')
        if dept_id:
            emp.departement_id = dept_id
        if manager_id:
            emp.manager_id = manager_id
        emp.save()
        _creer_soldes_employe(emp)
        messages.success(request, f"L'employé {emp.get_full_name()} a été créé avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la création : {e}")
    return redirect('/administration/?onglet=employes')


def _creer_soldes_employe(employe):
    annee = date.today().year
    for type_conge in TypeConge.objects.filter(actif=True):
        SoldeConge.objects.get_or_create(
            employe=employe,
            type_conge=type_conge,
            annee=annee,
            defaults={'jours_acquis': type_conge.jours_max}
        )


def _ajouter_type_conge(request):
    try:
        TypeConge.objects.create(
            code=request.POST.get('code', ''),
            libelle=request.POST.get('libelle', ''),
            couleur=request.POST.get('couleur', '#2196F3'),
            jours_max=int(request.POST.get('jours_max', 30)),
            necessite_justificatif=request.POST.get('necessite_justificatif') == 'on',
            description=request.POST.get('description', ''),
        )
        messages.success(request, "Type de congé ajouté avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur : {e}")
    return redirect('/administration/?onglet=types_conge')


def _initialiser_soldes(request):
    annee = int(request.POST.get('annee', date.today().year))
    count = 0
    for emp in Employe.objects.filter(actif=True):
        for tc in TypeConge.objects.filter(actif=True):
            _, created = SoldeConge.objects.get_or_create(
                employe=emp, type_conge=tc, annee=annee,
                defaults={'jours_acquis': tc.jours_max}
            )
            if created:
                count += 1
    messages.success(request, f"{count} soldes initialisés pour {annee}.")
    return redirect('/administration/?onglet=soldes')


@login_required
def importer_employes(request):
    if request.user.role not in ('rh', 'admin'):
        messages.error(request, "Accès non autorisé.")
        return redirect('administration')

    if request.method != 'POST' or not request.FILES.get('fichier_excel'):
        messages.error(request, "Aucun fichier fourni.")
        return redirect('/administration/?onglet=employes')

    import openpyxl
    fichier = request.FILES['fichier_excel']
    try:
        wb = openpyxl.load_workbook(fichier)
        ws = wb.active
    except Exception:
        messages.error(request, "Fichier Excel invalide.")
        return redirect('/administration/?onglet=employes')

    headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
    required = {'prenom', 'nom', 'email'}
    if not required.issubset(set(headers)):
        messages.error(request, "Colonnes obligatoires manquantes : Prenom, Nom, Email.")
        return redirect('/administration/?onglet=employes')

    def col(row, name):
        try:
            idx = headers.index(name)
            val = row[idx].value
            return str(val).strip() if val is not None else ''
        except (ValueError, IndexError):
            return ''

    crees, mis_a_jour, erreurs = 0, 0, []
    annee = date.today().year
    type_annuel = TypeConge.objects.filter(code='annuel').first()

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if not any(c.value for c in row):
            continue
        prenom = col(row, 'prenom')
        nom    = col(row, 'nom')
        email  = col(row, 'email')
        if not prenom or not nom or not email:
            erreurs.append(f"Ligne {i} : prénom, nom ou email manquant.")
            continue

        matricule   = col(row, 'matricule')
        poste       = col(row, 'poste')
        role        = col(row, 'role') or 'employe'
        dept_code   = col(row, 'departement')
        manager_email = col(row, 'manager')
        password    = col(row, 'mot de passe') or 'Petrosen2025!'

        if role not in ('employe', 'manager', 'rh', 'admin'):
            role = 'employe'

        dept = None
        if dept_code:
            dept = Departement.objects.filter(
                Q(code__iexact=dept_code) | Q(nom__iexact=dept_code)
            ).first()

        manager = None
        if manager_email:
            manager = Employe.objects.filter(email=manager_email).first()

        try:
            if Employe.objects.filter(email=email).exists():
                emp = Employe.objects.get(email=email)
                emp.first_name = prenom
                emp.last_name  = nom
                emp.poste      = poste
                emp.role       = role
                if dept:
                    emp.departement = dept
                if manager:
                    emp.manager = manager
                if matricule:
                    emp.matricule = matricule
                emp.actif = True
                emp.save()
                mis_a_jour += 1
            else:
                username = email.split('@')[0]
                if Employe.objects.filter(username=username).exists():
                    username = email.replace('@', '_').replace('.', '_')
                emp = Employe.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=prenom,
                    last_name=nom,
                )
                emp.poste       = poste
                emp.role        = role
                emp.departement = dept
                emp.manager     = manager
                emp.matricule   = matricule
                emp.actif       = True
                emp.save()
                if type_annuel:
                    SoldeConge.objects.get_or_create(
                        employe=emp, type_conge=type_annuel, annee=annee,
                        defaults={'jours_acquis': 24}
                    )
                crees += 1
        except Exception as e:
            erreurs.append(f"Ligne {i} ({email}) : {e}")

    msg = f"Import terminé : {crees} créé(s), {mis_a_jour} mis à jour."
    if erreurs:
        msg += f" {len(erreurs)} erreur(s) : " + " | ".join(erreurs[:5])
        messages.warning(request, msg)
    else:
        messages.success(request, msg)

    return redirect('/administration/?onglet=employes')


@login_required
def telecharger_template_employes(request):
    if request.user.role not in ('rh', 'admin'):
        messages.error(request, "Accès non autorisé.")
        return redirect('administration')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employés"

    headers = ['Prenom', 'Nom', 'Email', 'Matricule', 'Poste',
               'Departement', 'Role', 'Manager', 'Mot de passe']
    widths  = [15, 15, 30, 12, 30, 30, 12, 30, 18]

    header_fill = PatternFill("solid", fgColor="1B5E20")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = w

    exemples = [
        ['Aminata', 'Diallo', 'a.diallo@petrosen.sn', 'PET-001', 'Ingénieure Process',
         'DEP', 'employe', 'i.diop@petrosen.sn', 'Petrosen2025!'],
        ['Moussa', 'Ndiaye', 'm.ndiaye@petrosen.sn', 'PET-002', 'Comptable',
         'FIN', 'employe', 'o.kane@petrosen.sn', 'Petrosen2025!'],
    ]
    for r_idx, row in enumerate(exemples, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Feuille aide
    ws2 = wb.create_sheet("Aide")
    ws2['A1'] = "Valeurs acceptées pour la colonne Role :"
    ws2['A2'] = "employe  |  manager  |  rh  |  admin"
    ws2['A4'] = "Colonne Departement : code (DG, DRH, DEP, FIN, HSE, JUR, DSI, LOG) ou nom complet"
    ws2['A6'] = "Colonne Manager : email du manager direct (doit exister dans le système)"
    ws2['A8'] = "Mot de passe : laissez vide pour utiliser Petrosen2025! par défaut"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modele_employes_petrosen.xlsx"'
    wb.save(response)
    return response


@login_required
def notifications(request):
    user = request.user
    notifs = user.notifications.all()[:50]
    user.notifications.filter(lue=False).update(lue=True)
    context = {
        'notifs': notifs,
        'notifications_non_lues': 0,
    }
    return render(request, 'conges/notifications.html', context)


@login_required
def api_notifications(request):
    notifs = request.user.notifications.filter(lue=False)[:10]
    data = [{
        'id': n.id,
        'titre': n.titre,
        'message': n.message,
        'lien': n.lien,
        'date': n.date_creation.strftime('%d/%m/%Y %H:%M'),
    } for n in notifs]
    return JsonResponse({'notifications': data, 'count': len(data)})


@login_required
def profil(request):
    user = request.user
    if request.method == 'POST':
        user.first_name = request.POST.get('prenom', user.first_name)
        user.last_name = request.POST.get('nom', user.last_name)
        user.telephone = request.POST.get('telephone', user.telephone)
        if request.FILES.get('avatar'):
            user.avatar = request.FILES['avatar']
        new_pwd = request.POST.get('nouveau_mot_de_passe')
        if new_pwd:
            current_pwd = request.POST.get('mot_de_passe_actuel')
            if user.check_password(current_pwd):
                user.set_password(new_pwd)
                messages.success(request, "Mot de passe modifié.")
            else:
                messages.error(request, "Mot de passe actuel incorrect.")
        user.save()
        messages.success(request, "Profil mis à jour.")
        return redirect('profil')

    context = {
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/profil.html', context)
